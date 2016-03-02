# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
__author__ = "Cydonia"

# wb = load_workbook(filename="testlist.xlsx")
# ws = wb['Sheet1']
# runlist = []
#
# for row in ws.rows:
#     if row[1].value == 'Y':
#         runlist.append(row[0].value)
#
# print(runlist)

flag = 'passed'
msg = 'tested succeed'
test_result = [flag, msg]

wb = Workbook()
ws = wb.create_sheet(title='result')
for row in range(1, 10):
    ws.append(test_result)

wb.save('test_reuslt.xlsx')